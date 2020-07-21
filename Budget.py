import requests
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import openpyxl
from operator import itemgetter

theFile = openpyxl.load_workbook('Bollywood_masterdata_1.xlsx')
print(theFile.sheetnames)
currentSheet = theFile['Sheet1']

Movies = []
Budgets = []

#for row in range(2, currentSheet.max_row + 1):
for row in range(301, 314):
    for column in "B":  # Here you can add or reduce the columns
        cell_name = "{}{}".format(column, row)
        movie_name = currentSheet[cell_name].value
        movie = currentSheet[cell_name].value.replace(" ", "-")
        print(cell_name)
        b = 'https://bestoftheyear.in/movie/'+movie+'/'
        print(b)

        try:
            site = requests.get(b)
        except ValueError:
            pass
        soup = BeautifulSoup(site.content, 'html.parser')

        # Get all the tables
        table = soup.find('table', id="movie-info")
        #print(table)

        try:
            labels = table.find_all('th', class_="cell")
            for label in labels:
                if label.text == "Budget":
                    i = labels.index(label)
                    #print(i)
                else:
                    continue

            #print(labels)
            values = table.find_all('td', class_="value")
            Budgets.append(values[i].text)
            Movies.append(movie_name)
        except AttributeError:
                continue

#print(Movies)
#print(Budgets)

test_df = pd.DataFrame({'movie': Movies,
'budget': Budgets
})
print(test_df.info())
test_df
test_df.to_csv("bestoftheyear budget-3.csv", sep=',', index=False)
